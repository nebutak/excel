from __future__ import annotations

import re
import unicodedata
from pathlib import Path

from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from config import RuntimeConfig
from models import SheetMetadata, TableLayout

NUMBER_PATTERN = re.compile(r"-?\d+(?:[\.,]\d+)?")

HEADER_ALIASES: dict[str, tuple[str, ...]] = {
    "depth": ("do sau",),
    "instant_flow": ("luu luong tuc thi",),
    "flow_segment": ("luu luong doan",),
    "main_hoist_speed": ("toc do toi chinh",),
}

METADATA_LABELS: dict[str, tuple[str, ...]] = {
    "pile_no": ("so coc",),
    "target_depth": ("do sau muc tieu",),
    "diameter": ("duong kinh coc",),
    "header_drill_depth": ("khoan sau",),
    "accumulated_flow_raw": ("luu luong tich luy",),
    "start_time_raw": ("thoi gian bat dau",),
    "end_time_raw": ("thoi gian ket thuc",),
}


def parse_number(value: object) -> float | None:
    if value is None or value == "":
        return None
    if isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    match = NUMBER_PATTERN.search(str(value))
    if not match:
        return None
    return float(match.group(0).replace(",", "."))


def normalize_text(value: object) -> str:
    if value is None:
        return ""
    text = str(value).strip().lower()
    text = text.replace("đ", "d")
    text = unicodedata.normalize("NFKD", text)
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    text = re.sub(r"[^a-z0-9]+", " ", text)
    return " ".join(text.split())


def safe_average(values: list[float], exclude_zero: bool = False) -> float | None:
    filtered = [value for value in values if not exclude_zero or abs(value) > 1e-9]
    if not filtered:
        return None
    return sum(filtered) / len(filtered)


def build_range(column_index: int, start_row: int, end_row: int) -> str:
    return f"{get_column_letter(column_index)}{start_row}:{get_column_letter(column_index)}{end_row}"


def is_sheet_effectively_empty(worksheet: Worksheet) -> bool:
    for row in worksheet.iter_rows():
        for cell in row:
            if cell.value not in (None, ""):
                return False
    return True


def find_header_row(worksheet: Worksheet, max_scan_rows: int = 30) -> int | None:
    best_row: int | None = None
    best_score = -1
    for row_idx in range(1, min(max_scan_rows, worksheet.max_row) + 1):
        normalized_cells = [normalize_text(worksheet.cell(row_idx, col_idx).value) for col_idx in range(1, worksheet.max_column + 1)]
        score = 0
        for aliases in HEADER_ALIASES.values():
            if any(any(alias in cell_value for alias in aliases) for cell_value in normalized_cells):
                score += 1
        if score > best_score:
            best_score = score
            best_row = row_idx
    if best_score < 3:
        return None
    return best_row


def _match_alias(cell_value: str, aliases: tuple[str, ...]) -> bool:
    return any(alias in cell_value for alias in aliases)


def detect_table_layout(worksheet: Worksheet, config: RuntimeConfig) -> TableLayout:
    header_row = find_header_row(worksheet)
    if header_row is None:
        return TableLayout(
            header_row=10,
            depth_col=column_index_from_string(config.fallback_columns["depth"]),
            instant_flow_col=column_index_from_string(config.fallback_columns["instant_flow"]),
            flow_col=column_index_from_string(config.fallback_columns["flow_segment"]),
            speed_col=column_index_from_string(config.fallback_columns["main_hoist_speed"]),
            last_data_col=13,
            used_fallback=True,
        )

    matched: dict[str, int] = {}
    last_data_col = 13
    for col_idx in range(1, worksheet.max_column + 1):
        normalized_value = normalize_text(worksheet.cell(header_row, col_idx).value)
        if not normalized_value:
            continue
        last_data_col = max(last_data_col, col_idx)
        for field_name, aliases in HEADER_ALIASES.items():
            if field_name not in matched and _match_alias(normalized_value, aliases):
                matched[field_name] = col_idx

    required_fields = {"depth", "flow_segment", "main_hoist_speed"}
    if required_fields.issubset(matched):
        return TableLayout(
            header_row=header_row,
            depth_col=matched["depth"],
            instant_flow_col=matched.get("instant_flow", column_index_from_string(config.fallback_columns["instant_flow"])),
            flow_col=matched["flow_segment"],
            speed_col=matched["main_hoist_speed"],
            last_data_col=last_data_col,
            used_fallback=False,
        )

    return TableLayout(
        header_row=header_row,
        depth_col=matched.get("depth", column_index_from_string(config.fallback_columns["depth"])),
        instant_flow_col=matched.get("instant_flow", column_index_from_string(config.fallback_columns["instant_flow"])),
        flow_col=matched.get("flow_segment", column_index_from_string(config.fallback_columns["flow_segment"])),
        speed_col=matched.get("main_hoist_speed", column_index_from_string(config.fallback_columns["main_hoist_speed"])),
        last_data_col=last_data_col,
        used_fallback=True,
    )


def extract_metadata(worksheet: Worksheet) -> SheetMetadata:
    metadata = SheetMetadata()
    lookup: dict[str, tuple[int, int]] = {}
    max_row = min(10, worksheet.max_row)
    max_col = min(8, worksheet.max_column)

    for row_idx in range(1, max_row + 1):
        for col_idx in range(1, max_col + 1):
            normalized = normalize_text(worksheet.cell(row_idx, col_idx).value)
            if not normalized:
                continue
            for field_name, aliases in METADATA_LABELS.items():
                if field_name in lookup:
                    continue
                if _match_alias(normalized, aliases):
                    lookup[field_name] = (row_idx, col_idx)

    for field_name, (row_idx, col_idx) in lookup.items():
        right_value = worksheet.cell(row_idx, col_idx + 1).value
        far_right_value = worksheet.cell(row_idx, col_idx + 2).value
        if field_name in {"start_time_raw", "end_time_raw"}:
            parts = [str(part).strip() for part in (right_value, far_right_value) if part not in (None, "")]
            setattr(metadata, field_name, " ".join(parts) if parts else None)
            continue
        if field_name == "pile_no":
            setattr(metadata, field_name, None if right_value in (None, "") else str(right_value).strip())
            continue
        raw_value = None if right_value in (None, "") else str(right_value).strip()
        if field_name == "target_depth":
            metadata.target_depth = parse_number(right_value)
            metadata.target_depth_raw = raw_value
        elif field_name == "diameter":
            metadata.diameter = parse_number(right_value)
            metadata.diameter_raw = raw_value
        elif field_name == "header_drill_depth":
            metadata.header_drill_depth = parse_number(right_value)
            metadata.header_drill_depth_raw = raw_value
        elif field_name == "accumulated_flow_raw":
            metadata.accumulated_flow_raw = raw_value

    return metadata


def find_result_anchor(
    worksheet: Worksheet,
    start_row: int,
    table_end_col: int,
) -> tuple[int, int]:
    label_candidates = {"cd dinh", "llv", "vx", "vl"}
    for row_idx in range(max(1, start_row - 2), min(worksheet.max_row, start_row + 12) + 1):
        for col_idx in range(14, max(20, worksheet.max_column) + 1):
            normalized = normalize_text(worksheet.cell(row_idx, col_idx).value)
            if normalized in label_candidates:
                return row_idx, col_idx

    candidate_col = max(table_end_col + 1, 14)
    while candidate_col < 50:
        blocked = False
        for row_idx in range(start_row, min(worksheet.max_row, start_row + 8) + 1):
            if worksheet.cell(row_idx, candidate_col).value not in (None, ""):
                blocked = True
                break
            if worksheet.cell(row_idx, candidate_col + 1).value not in (None, ""):
                blocked = True
                break
        if not blocked:
            return start_row, candidate_col
        candidate_col += 1
    return start_row, max(table_end_col + 2, 15)


def make_output_path(
    input_path: Path,
    output_dir: Path,
    suffix: str,
    avoid_overwrite: bool = True,
) -> Path:
    output_dir.mkdir(parents=True, exist_ok=True)
    base_name = f"{input_path.stem}{suffix}{input_path.suffix}"
    output_path = output_dir / base_name
    if not avoid_overwrite:
        return output_path

    counter = 1
    while output_path.exists():
        output_path = output_dir / f"{input_path.stem}{suffix}_{counter}{input_path.suffix}"
        counter += 1
    return output_path
