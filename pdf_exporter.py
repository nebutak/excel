from __future__ import annotations

import shutil
import subprocess
import tempfile
from pathlib import Path
from typing import Callable

from openpyxl import load_workbook

LogCallback = Callable[[str], None] | None
ProgressCallback = Callable[[int, str], None] | None

PDF_FORMAT = 0
INVALID_FILENAME_CHARS = '<>:"/\\|?*'


class PdfExportError(RuntimeError):
    pass


def list_sheet_names(input_path: str | Path) -> list[str]:
    workbook = load_workbook(input_path, read_only=True)
    try:
        return list(workbook.sheetnames)
    finally:
        workbook.close()


def export_sheets_to_pdf(
    input_path: str | Path,
    sheet_names: list[str],
    output_dir: str | Path,
    progress_callback: ProgressCallback = None,
    log_callback: LogCallback = None,
) -> list[Path]:
    input_path = Path(input_path)
    output_dir = Path(output_dir)
    if not sheet_names:
        raise PdfExportError("Chưa chọn sheet để xuất PDF.")
    output_dir.mkdir(parents=True, exist_ok=True)

    try:
        return _export_with_excel_com(input_path, sheet_names, output_dir, progress_callback, log_callback)
    except PdfExportError as excel_error:
        soffice_path = _find_soffice()
        if not soffice_path:
            raise PdfExportError(
                f"{excel_error} Không tìm thấy LibreOffice để xuất PDF dự phòng."
            ) from excel_error
        _log(log_callback, "Không dùng được Excel COM, chuyển sang LibreOffice headless.")
        return _export_with_libreoffice(
            input_path,
            sheet_names,
            output_dir,
            soffice_path,
            progress_callback,
            log_callback,
        )


def preview_sheet(input_path: str | Path, sheet_name: str) -> None:
    try:
        import win32com.client  # type: ignore[import-not-found]
    except ImportError as exc:
        raise PdfExportError("Chức năng xem trước cần Microsoft Excel và pywin32 trên Windows.") from exc

    excel = win32com.client.DispatchEx("Excel.Application")
    workbook = None
    try:
        excel.Visible = True
        excel.DisplayAlerts = False
        workbook = excel.Workbooks.Open(str(Path(input_path).resolve()))
        worksheet = workbook.Worksheets(sheet_name)
        worksheet.Activate()
        worksheet.PrintPreview()
    finally:
        if workbook is not None:
            workbook.Close(SaveChanges=False)
        excel.Quit()


def print_sheets(
    input_path: str | Path,
    sheet_names: list[str],
    progress_callback: ProgressCallback = None,
    log_callback: LogCallback = None,
) -> None:
    if not sheet_names:
        raise PdfExportError("Chưa chọn sheet để in.")
    try:
        import win32com.client  # type: ignore[import-not-found]
    except ImportError as exc:
        raise PdfExportError("Chức năng in trực tiếp cần Microsoft Excel và pywin32 trên Windows.") from exc

    excel = None
    workbook = None
    try:
        excel = win32com.client.DispatchEx("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False
        workbook = excel.Workbooks.Open(str(Path(input_path).resolve()))
        total = len(sheet_names)
        for index, sheet_name in enumerate(sheet_names, start=1):
            _progress(progress_callback, index - 1, total, f"Đang gửi lệnh in sheet {sheet_name}...")
            worksheet = workbook.Worksheets(sheet_name)
            worksheet.PrintOut()
            _log(log_callback, f"Đã gửi sheet tới máy in: {sheet_name}")
            _progress(progress_callback, index, total, f"Đã gửi lệnh in sheet {sheet_name}")
    except Exception as exc:  # noqa: BLE001
        raise PdfExportError(f"Lỗi in trực tiếp bằng Excel: {exc}") from exc
    finally:
        if workbook is not None:
            workbook.Close(SaveChanges=False)
        if excel is not None:
            excel.Quit()


def _export_with_excel_com(
    input_path: Path,
    sheet_names: list[str],
    output_dir: Path,
    progress_callback: ProgressCallback,
    log_callback: LogCallback,
) -> list[Path]:
    try:
        import win32com.client  # type: ignore[import-not-found]
    except ImportError as exc:
        raise PdfExportError("Excel COM không khả dụng. Cần cài pywin32 trên Windows.") from exc

    excel = None
    workbook = None
    outputs: list[Path] = []
    try:
        excel = win32com.client.DispatchEx("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False
        workbook = excel.Workbooks.Open(str(input_path.resolve()))
        total = len(sheet_names)
        for index, sheet_name in enumerate(sheet_names, start=1):
            _progress(progress_callback, index - 1, total, f"Đang xuất PDF sheet {sheet_name}...")
            worksheet = workbook.Worksheets(sheet_name)
            output_path = _unique_output_path(output_dir, sheet_name)
            worksheet.ExportAsFixedFormat(PDF_FORMAT, str(output_path.resolve()))
            outputs.append(output_path)
            _log(log_callback, f"Đã xuất PDF: {output_path}")
            _progress(progress_callback, index, total, f"Đã xuất PDF sheet {sheet_name}")
    except Exception as exc:  # noqa: BLE001
        raise PdfExportError(f"Lỗi xuất PDF bằng Excel: {exc}") from exc
    finally:
        if workbook is not None:
            workbook.Close(SaveChanges=False)
        if excel is not None:
            excel.Quit()
    return outputs


def _export_with_libreoffice(
    input_path: Path,
    sheet_names: list[str],
    output_dir: Path,
    soffice_path: str,
    progress_callback: ProgressCallback,
    log_callback: LogCallback,
) -> list[Path]:
    outputs: list[Path] = []
    total = len(sheet_names)
    with tempfile.TemporaryDirectory() as temp_dir_name:
        temp_dir = Path(temp_dir_name)
        for index, sheet_name in enumerate(sheet_names, start=1):
            _progress(progress_callback, index - 1, total, f"Đang xuất PDF sheet {sheet_name}...")
            temp_xlsx = temp_dir / f"{_safe_filename(sheet_name)}.xlsx"
            _save_single_sheet_workbook(input_path, sheet_name, temp_xlsx)
            completed = subprocess.run(
                [
                    soffice_path,
                    "--headless",
                    "--convert-to",
                    "pdf",
                    "--outdir",
                    str(temp_dir),
                    str(temp_xlsx),
                ],
                check=False,
                capture_output=True,
                text=True,
            )
            if completed.returncode != 0:
                raise PdfExportError(completed.stderr.strip() or completed.stdout.strip() or "LibreOffice xuất PDF thất bại.")
            generated_pdf = temp_xlsx.with_suffix(".pdf")
            output_path = _unique_output_path(output_dir, sheet_name)
            shutil.move(str(generated_pdf), output_path)
            outputs.append(output_path)
            _log(log_callback, f"Đã xuất PDF: {output_path}")
            _progress(progress_callback, index, total, f"Đã xuất PDF sheet {sheet_name}")
    return outputs


def _save_single_sheet_workbook(input_path: Path, sheet_name: str, temp_xlsx: Path) -> None:
    workbook = load_workbook(input_path)
    try:
        if sheet_name not in workbook.sheetnames:
            raise PdfExportError(f"Không tìm thấy sheet: {sheet_name}")
        for name in list(workbook.sheetnames):
            if name != sheet_name:
                del workbook[name]
        workbook.active = 0
        workbook.save(temp_xlsx)
    finally:
        workbook.close()


def _find_soffice() -> str | None:
    for executable in ("soffice", "libreoffice"):
        found = shutil.which(executable)
        if found:
            return found
    return None


def _unique_output_path(output_dir: Path, sheet_name: str) -> Path:
    stem = _safe_filename(sheet_name)
    output_path = output_dir / f"{stem}.pdf"
    counter = 1
    while output_path.exists():
        output_path = output_dir / f"{stem}_{counter}.pdf"
        counter += 1
    return output_path


def _safe_filename(sheet_name: str) -> str:
    cleaned = "".join("_" if char in INVALID_FILENAME_CHARS else char for char in sheet_name).strip()
    return cleaned or "sheet"


def _log(callback: LogCallback, message: str) -> None:
    if callback:
        callback(message)


def _progress(callback: ProgressCallback, completed: int, total: int, message: str) -> None:
    if callback:
        callback(int((completed / max(total, 1)) * 100), message)
