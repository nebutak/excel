from __future__ import annotations

import csv
import re
from pathlib import Path
from typing import Callable

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from calculator import CalculationError, calculate_sheet
from config import SUMMARY_HEADERS, RuntimeConfig, clone_config
from models import SheetResult, TableRow
from utils import (
    build_range,
    detect_table_layout,
    extract_metadata,
    find_result_anchor,
    is_sheet_effectively_empty,
    make_output_path,
    normalize_text,
    parse_number,
)

LogCallback = Callable[[str], None] | None
ProgressCallback = Callable[[int, str], None] | None
NUMERIC_CELL_PATTERN = re.compile(r"^-?\d+(?:[\.,]\d+)?$")


class ExcelProcessor:
    def __init__(self, config: RuntimeConfig | None = None) -> None:
        self.config = clone_config(config)

    def process_workbook(
        self,
        input_path: str | Path,
        output_dir: str | Path,
        drill_length: float,
        progress_callback: ProgressCallback = None,
        log_callback: LogCallback = None,
    ) -> tuple[Path, list[SheetResult]]:
        input_path = Path(input_path)
        output_dir = Path(output_dir)
        workbook = load_workbook(input_path)
        results: list[SheetResult] = []

        if self.config.create_summary_sheet and "Tong_hop" in workbook.sheetnames:
            del workbook["Tong_hop"]

        worksheets = list(workbook.worksheets)
        total_sheets = max(len(worksheets), 1)

        for index, worksheet in enumerate(worksheets, start=1):
            self._report_progress(progress_callback, index - 1, total_sheets, f"Đang xử lý {worksheet.title}...")
            result = self._process_sheet(worksheet, drill_length, log_callback)
            results.append(result)
            self._report_progress(progress_callback, index, total_sheets, f"Đã xử lý {worksheet.title}")

        if self.config.create_summary_sheet:
            self._write_summary_sheet(workbook, results)

        output_path = make_output_path(
            input_path=input_path,
            output_dir=output_dir,
            suffix=self.config.output_suffix,
            avoid_overwrite=self.config.do_not_overwrite_original,
        )
        workbook.save(output_path)
        if log_callback:
            log_callback(f"Đã lưu file kết quả: {output_path}")
        return output_path, results

    def process_csv_files(
        self,
        csv_paths: list[str | Path],
        output_dir: str | Path,
        drill_length: float,
        progress_callback: ProgressCallback = None,
        log_callback: LogCallback = None,
    ) -> tuple[Path, list[SheetResult]]:
        paths = [Path(path) for path in csv_paths]
        if not paths:
            raise ValueError("Chưa chọn file CSV.")

        output_dir = Path(output_dir)
        workbook = Workbook()
        workbook.remove(workbook.active)
        results: list[SheetResult] = []
        total_files = max(len(paths), 1)

        for index, csv_path in enumerate(paths, start=1):
            self._report_progress(progress_callback, index - 1, total_files, f"Đang đọc {csv_path.name}...")
            rows = self._read_csv_rows(csv_path)
            sheet_title = self._build_csv_sheet_title(csv_path, rows, workbook.sheetnames)
            worksheet = workbook.create_sheet(sheet_title)
            self._write_csv_rows_to_sheet(worksheet, rows)
            result = self._process_sheet(worksheet, drill_length, log_callback)
            results.append(result)
            self._report_progress(progress_callback, index, total_files, f"Đã xử lý {csv_path.name}")

        if self.config.create_summary_sheet:
            self._write_summary_sheet(workbook, results)

        base_stem = paths[0].stem if len(paths) == 1 else "csv_tong_hop"
        output_path = self._make_xlsx_output_path(output_dir, base_stem)
        workbook.save(output_path)
        if log_callback:
            log_callback(f"Đã lưu file kết quả: {output_path}")
        return output_path, results

    def _read_csv_rows(self, csv_path: Path) -> list[list[str]]:
        last_decode_error: UnicodeDecodeError | None = None
        for encoding in ("utf-8-sig", "utf-8", "cp1258", "cp1252"):
            try:
                with csv_path.open("r", encoding=encoding, newline="") as file_obj:
                    sample = file_obj.read(4096)
                    file_obj.seek(0)
                    try:
                        dialect = csv.Sniffer().sniff(sample, delimiters=",;\t")
                    except csv.Error:
                        dialect = csv.excel
                    return [row for row in csv.reader(file_obj, dialect)]
            except UnicodeDecodeError as exc:
                last_decode_error = exc
        raise ValueError(f"Không đọc được encoding CSV: {csv_path.name}") from last_decode_error

    def _write_csv_rows_to_sheet(self, worksheet: Worksheet, rows: list[list[str]]) -> None:
        for row_idx, row in enumerate(rows, start=1):
            for col_idx, value in enumerate(row, start=1):
                worksheet.cell(row_idx, col_idx, self._coerce_csv_cell_value(value))

    def _coerce_csv_cell_value(self, value: str) -> str | int | float | None:
        stripped = value.strip()
        if not stripped:
            return None
        if not NUMERIC_CELL_PATTERN.fullmatch(stripped):
            return value
        normalized = stripped.replace(",", ".")
        number_value = float(normalized)
        if number_value.is_integer() and "." not in normalized:
            return int(number_value)
        return number_value

    def _build_csv_sheet_title(
        self,
        csv_path: Path,
        rows: list[list[str]],
        existing_titles: list[str],
    ) -> str:
        pile_no = self._extract_csv_pile_no(rows)
        base_title = pile_no or csv_path.stem
        cleaned = "".join("_" if char in r'[]:*?/\\' else char for char in base_title).strip()
        cleaned = cleaned or "CSV"
        cleaned = cleaned[:31]

        title = cleaned
        counter = 1
        while title in existing_titles:
            suffix = f"_{counter}"
            title = f"{cleaned[:31 - len(suffix)]}{suffix}"
            counter += 1
        return title

    def _extract_csv_pile_no(self, rows: list[list[str]]) -> str | None:
        for row in rows[:10]:
            for col_idx, value in enumerate(row):
                if normalize_text(value) == "so coc" and col_idx + 1 < len(row):
                    pile_value = row[col_idx + 1]
                    if pile_value not in (None, ""):
                        return str(pile_value).strip()
        return None

    def _make_xlsx_output_path(self, output_dir: Path, base_stem: str) -> Path:
        output_dir.mkdir(parents=True, exist_ok=True)
        output_path = output_dir / f"{base_stem}{self.config.output_suffix}.xlsx"
        if not self.config.do_not_overwrite_original:
            return output_path

        counter = 1
        while output_path.exists():
            output_path = output_dir / f"{base_stem}{self.config.output_suffix}_{counter}.xlsx"
            counter += 1
        return output_path

    def _process_sheet(
        self,
        worksheet: Worksheet,
        drill_length: float,
        log_callback: LogCallback,
    ) -> SheetResult:
        if is_sheet_effectively_empty(worksheet):
            note = "Sheet trống, bỏ qua."
            self._log(log_callback, f"{worksheet.title}: {note}")
            return SheetResult(
                sheet_name=worksheet.title,
                pile_no=None,
                target_depth=None,
                header_drill_depth=None,
                actual_drill_depth=None,
                drill_length=drill_length,
                design_top_elevation=self.config.default_design_top_elevation,
                design_bottom_elevation=self.config.default_design_bottom_elevation,
                cd_dinh=None,
                llv=None,
                vx=None,
                vl=None,
                actual_top_elevation=None,
                actual_bottom_elevation=None,
                actual_column_length=None,
                vx_range=None,
                vl_range=None,
                llv_range=None,
                status="SKIP",
                note=note,
            )

        metadata = extract_metadata(worksheet)
        layout = detect_table_layout(worksheet, self.config)
        notes: list[str] = []
        if layout.used_fallback:
            notes.append("Không dò đủ header, đã dùng fallback cột C/F/H.")

        rows = self._read_table_rows(worksheet, layout)
        if not rows:
            note = "Không tìm thấy bảng dữ liệu hợp lệ."
            self._log(log_callback, f"{worksheet.title}: {note}")
            return self._build_error_result(worksheet.title, metadata, drill_length, note)

        if metadata.pile_no is None:
            for row in rows:
                pile_value = worksheet.cell(row.excel_row, 12).value
                if pile_value not in (None, ""):
                    metadata.pile_no = str(pile_value).strip()
                    break

        try:
            calculation = calculate_sheet(rows, metadata.target_depth, drill_length, self.config)
        except CalculationError as exc:
            note = str(exc)
            self._log(log_callback, f"{worksheet.title}: {note}")
            return self._build_error_result(
                worksheet.title,
                metadata,
                drill_length,
                note,
                actual_drill_depth=max((row.depth for row in rows), default=None),
            )

        notes.extend(calculation.notes)
        note = " | ".join(note_item for note_item in notes if note_item)
        status = "WARNING" if notes else "OK"

        vx_range = build_range(layout.speed_col, rows[calculation.vx_start_idx].excel_row, rows[calculation.vx_end_idx].excel_row)
        vl_range = build_range(layout.speed_col, rows[calculation.vl_start_idx].excel_row, rows[calculation.vl_end_idx].excel_row)
        llv_range = (
            build_range(layout.flow_col, rows[calculation.llv_start_idx].excel_row, rows[calculation.llv_end_idx].excel_row)
            if calculation.llv_end_idx >= calculation.llv_start_idx
            else ""
        )

        result = SheetResult(
            sheet_name=worksheet.title,
            pile_no=metadata.pile_no,
            target_depth=metadata.target_depth,
            header_drill_depth=metadata.header_drill_depth,
            actual_drill_depth=calculation.actual_drill_depth,
            drill_length=drill_length,
            design_top_elevation=self.config.default_design_top_elevation,
            design_bottom_elevation=self.config.default_design_bottom_elevation,
            cd_dinh=calculation.cd_dinh,
            llv=calculation.llv,
            vx=calculation.vx,
            vl=calculation.vl,
            actual_top_elevation=calculation.actual_top_elevation,
            actual_bottom_elevation=calculation.actual_bottom_elevation,
            actual_column_length=calculation.actual_column_length,
            vx_range=vx_range,
            vl_range=vl_range,
            llv_range=llv_range,
            status=status,
            note=note,
        )

        self._write_sheet_result(worksheet, layout, result)
        if self.config.highlight_ranges:
            self._highlight_ranges(worksheet, layout, rows, calculation)
        self._log(log_callback, f"{worksheet.title}: {status} - CD={result.cd_dinh:.2f}, LLV={result.llv:.2f}, VX={result.vx:.4f}, VL={result.vl:.4f}")
        return result

    def _read_table_rows(self, worksheet: Worksheet, layout) -> list[TableRow]:
        rows: list[TableRow] = []
        blank_streak = 0
        for row_idx in range(layout.header_row + 1, worksheet.max_row + 1):
            depth = parse_number(worksheet.cell(row_idx, layout.depth_col).value)
            if depth is None:
                blank_streak += 1
                if blank_streak >= 2 and rows:
                    break
                continue
            blank_streak = 0
            flow_value = parse_number(worksheet.cell(row_idx, layout.flow_col).value) or 0.0
            instant_flow_value = parse_number(worksheet.cell(row_idx, layout.instant_flow_col).value) or 0.0
            speed_value = parse_number(worksheet.cell(row_idx, layout.speed_col).value) or 0.0
            rows.append(
                TableRow(
                    excel_row=row_idx,
                    depth=depth,
                    instant_flow=instant_flow_value,
                    flow_segment=flow_value,
                    main_hoist_speed=speed_value,
                )
            )
        return rows

    def _write_sheet_result(self, worksheet: Worksheet, layout, result: SheetResult) -> None:
        start_row, start_col = find_result_anchor(worksheet, layout.header_row, layout.last_data_col)
        items = [
            (self.config.result_labels["cd_dinh"], result.cd_dinh),
            (self.config.result_labels["llv"], result.llv),
            (self.config.result_labels["vx"], result.vx),
            (self.config.result_labels["vl"], result.vl),
            (self.config.result_labels["actual_column_length"], result.actual_column_length),
            (self.config.result_labels["actual_top_elevation"], result.actual_top_elevation),
            (self.config.result_labels["actual_bottom_elevation"], result.actual_bottom_elevation),
            (self.config.result_labels["vx_range"], result.vx_range),
            (self.config.result_labels["vl_range"], result.vl_range),
            (self.config.result_labels["llv_range"], result.llv_range),
            (self.config.result_labels["status"], result.status),
            (self.config.result_labels["note"], result.note),
        ]

        for offset, (label, value) in enumerate(items):
            label_cell = worksheet.cell(start_row + offset, start_col)
            value_cell = worksheet.cell(start_row + offset, start_col + 1)
            label_cell.value = label
            label_cell.font = Font(bold=True)
            value_cell.value = value
            if isinstance(value, float):
                if label in {self.config.result_labels["cd_dinh"]}:
                    value_cell.number_format = "0.00"
                elif label in {self.config.result_labels["vx"], self.config.result_labels["vl"]}:
                    value_cell.number_format = "0.0000"
                else:
                    value_cell.number_format = "0.00"

    def _highlight_ranges(self, worksheet: Worksheet, layout, rows: list[TableRow], calculation) -> None:
        fills = {
            "vx": PatternFill(fill_type="solid", start_color=self.config.highlight_colors["vx"], end_color=self.config.highlight_colors["vx"]),
            "vl": PatternFill(fill_type="solid", start_color=self.config.highlight_colors["vl"], end_color=self.config.highlight_colors["vl"]),
            "llv": PatternFill(fill_type="solid", start_color=self.config.highlight_colors["llv"], end_color=self.config.highlight_colors["llv"]),
        }

        for idx in range(calculation.vx_start_idx, calculation.vx_end_idx + 1):
            worksheet.cell(rows[idx].excel_row, layout.speed_col).fill = fills["vx"]
        for idx in range(calculation.vl_start_idx, calculation.vl_end_idx + 1):
            worksheet.cell(rows[idx].excel_row, layout.speed_col).fill = fills["vl"]
        for idx in range(calculation.llv_start_idx, calculation.llv_end_idx + 1):
            worksheet.cell(rows[idx].excel_row, layout.flow_col).fill = fills["llv"]

    def _write_summary_sheet(self, workbook: Workbook, results: list[SheetResult]) -> None:
        sheet = workbook.create_sheet("Tong_hop")
        for col_idx, header in enumerate(SUMMARY_HEADERS, start=1):
            cell = sheet.cell(1, col_idx, header)
            cell.font = Font(bold=True)

        for row_idx, result in enumerate(results, start=2):
            values = [
                row_idx - 1,
                result.sheet_name,
                result.pile_no or "",
                result.target_depth,
                result.header_drill_depth,
                result.actual_drill_depth,
                result.drill_length,
                result.design_top_elevation,
                result.design_bottom_elevation,
                result.cd_dinh,
                result.llv,
                result.vx,
                result.vl,
                result.actual_top_elevation,
                result.actual_bottom_elevation,
                result.actual_column_length,
                result.vx_range or "",
                result.vl_range or "",
                result.llv_range or "",
                result.status,
                result.note,
            ]
            for col_idx, value in enumerate(values, start=1):
                sheet.cell(row_idx, col_idx, value)

        sheet.auto_filter.ref = f"A1:{get_column_letter(len(SUMMARY_HEADERS))}{sheet.max_row}"
        for column_cells in sheet.columns:
            width = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column_cells)
            sheet.column_dimensions[column_cells[0].column_letter].width = min(max(width + 2, 12), 40)

    def _build_error_result(
        self,
        sheet_name: str,
        metadata,
        drill_length: float,
        note: str,
        actual_drill_depth: float | None = None,
    ) -> SheetResult:
        return SheetResult(
            sheet_name=sheet_name,
            pile_no=metadata.pile_no,
            target_depth=metadata.target_depth,
            header_drill_depth=metadata.header_drill_depth,
            actual_drill_depth=actual_drill_depth,
            drill_length=drill_length,
            design_top_elevation=self.config.default_design_top_elevation,
            design_bottom_elevation=self.config.default_design_bottom_elevation,
            cd_dinh=None,
            llv=None,
            vx=None,
            vl=None,
            actual_top_elevation=None,
            actual_bottom_elevation=None,
            actual_column_length=None,
            vx_range=None,
            vl_range=None,
            llv_range=None,
            status="ERROR",
            note=note,
        )

    def _log(self, callback: LogCallback, message: str) -> None:
        if callback:
            callback(message)

    def _report_progress(
        self,
        callback: ProgressCallback,
        completed: int,
        total: int,
        message: str,
    ) -> None:
        if callback:
            callback(int((completed / max(total, 1)) * 100), message)
